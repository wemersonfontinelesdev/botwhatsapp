"""
import openpyxl
from urllib.parse import quote
#formatar a mensagem
import webbrowser #para abrir o navegador.
from time import sleep  #TEMPO
import pyautogui
import os

#Colocar um tempo para a pagina carregadar inicialmente
webbrowser.open('https://web.whatsapp.com/')
sleep(70)
workbook = openpyxl.load_workbook('clientes.xlsx')

#carregar exatamente a primeira pagina da tabela.
pagina_clientes = workbook['Sheet1']
#interar
for linha in pagina_clientes.iter_rows(min_row=2):
  nome=linha[0].value
  telefone=linha[1].value
  curso=linha[2].value
  #print(nome)
  #print(telefone)
  #print(curso)



  #criar linkes personalizados do WhatsApp para cada cliente.
  #https://web.whatsapp.com/send?phone=xxxxxx&text="sdsfsss"
mensagem = f'Olá{nome}, ficamos feliz em saber do seu interesse no curso de {curso}. Só passando pra informar que estamos com ótimas condições para você realizar sua matricula.'


try:
  link_mensagem_whatsapp= f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'

  #abrir o navegador
  webbrowser.open(link_mensagem_whatsapp)
  sleep(17)
  seta = pyautogui.locateCenterOnScreen('seta.png')
  sleep(9)
  pyautogui.click(seta[0],seta[1])
  sleep(9)
  pyautogui.hotkey('ctrl', 'w')
  sleep(9)
except:
  print(f'Não foi possivel enviar mensagem para nome {nome}')
  with open('erros.csv','a',newline='', encoding='utf-8') as arquivo:
    arquivo.write(f'{telefone},{nome}')


"""

"""
import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui

# Carregar a planilha
workbook = openpyxl.load_workbook('COELHO NETO - MA.xlsx')
pagina_clientes = workbook['Planilha1']

# Abrir o WhatsApp Web uma única vez
webbrowser.open('https://web.whatsapp.com/')
sleep(30)  # Tempo para o WhatsApp Web carregar

# Iterar sobre as linhas da planilha
for linha in pagina_clientes.iter_rows(min_row=2):
    nome = linha[2].value
    telefone = linha[3].value
    curso = linha[4].value

    # Criar a mensagem personalizada
    mensagem = f'Olá {nome}, ficamos felizes em saber do seu interesse no curso de {curso}. Só passando pra informar que estamos com ótimas condições para você realizar sua matrícula.'

    try:
        # Criar o link do WhatsApp com a mensagem
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'

        # Abrir o link no navegador
        webbrowser.open(link_mensagem_whatsapp)
        sleep(17)  # Tempo para carregar o link

        # Localizar e clicar na seta para enviar a mensagem
        seta = pyautogui.locateCenterOnScreen('WhatsApp Image 2024-08-06 at 16.52.40.jpeg')
        if seta is not None:
            pyautogui.click(seta[0], seta[1])
            sleep(7)
            pyautogui.hotkey('ctrl', 'w')  # Fechar a aba do WhatsApp Web
            sleep(7)
        else:
            print(f'Não foi possível localizar a imagem para {nome}.')

    except Exception as e:
        print(f'Não foi possível enviar mensagem para {nome}: {e}')
        with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{telefone},{nome}\n')

"""
import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import os

webbrowser.open('https://web.whatsapp.com/')
# Aumentar o tempo de espera inicial para garantir que o WhatsApp Web carregue completamente
sleep(60)  # Aguarda 60 segundos para o WhatsApp Web carregar

# Ler planilha e guardar informações sobre nome, telefone e data de vencimento
workbook = openpyxl.load_workbook('clientes.xlsx')
pagina_clientes = workbook['Sheet1']

for linha in pagina_clientes.iter_rows(min_row=2):
    nome = linha[0].value
    telefone = linha[1].value
    vencimento = linha[2].value

    mensagem = f"Olá {nome}, seu boleto vence no dia {vencimento.strftime('%d/%m/%Y')}. Favor pagar no link"

    try:
        telefone_formatado= f"+{telefone}"
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone_formatado}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)
        sleep(30)  # Aumentar o tempo de espera antes de tentar localizar o botão
        seta = pyautogui.locateCenterOnScreen('seta.png')

        if seta:
            pyautogui.click(seta[0], seta[1])
            sleep(7)
            pyautogui.hotkey('ctrl', 'w')
        else:
            raise Exception("Botão não encontrado.")

        sleep(7)  # Espera antes de enviar a próxima mensagem
    except Exception as e:
        print(f'Não foi possível enviar mensagem para {nome}: {e}')
        with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome},{telefone}{os.linesep}')
